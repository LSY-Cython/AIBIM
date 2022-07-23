import requests
import json
import re
import xlwt
from property_filter import property_filter
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
import time
import os

# 查询符合条件的构件ID列表
def query_info(access_token, file_id, query_input):
    query_url = "https://api.bimface.com/data/v2/query/elementIds"
    query_header = {"Authorization": f"Bearer {access_token}"}  # 在请求头中添加上Authorization这个参数
    query_body = {"targetType": "file", "targetIds": [file_id],
                  "query": {"contain": {"family": query_input}}}  # 查询请求体字典要序列化成JSON字符串
    query_id = requests.post(url=query_url, headers=query_header, data=json.dumps(query_body)).json()['data'][0]['elementIds']
    return query_id

# 获取指定ID构件的属性
def get_info(access_token, file_id, element_id):
    element_url = f"https://api.bimface.com/data/v2/files/{file_id}/elements/{element_id}"
    element_header = {"Authorization": f"Bearer {access_token}"}
    element_info = requests.get(url=element_url, headers=element_header).json()
    return element_info

def check_info(access_token, file_id, socket_io, event):
    with open("formular.txt", 'r', encoding='utf-8') as f:
        formular_list = f.read().split("\n")
    row_index = 2
    # workbook = xlwt.Workbook(encoding='utf-8')
    # worksheet = workbook.add_sheet("审查结果")
    # worksheet = create_header(worksheet)
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(title="审查结果", index=0)
    worksheet = create_header(worksheet)
    for formular in formular_list:
        match = re.match(r'^\((.*?),(.*?)\)(.*?)$', formular)
        query_family = match.group(1)
        property_name = match.group(2).split("/")  # 比值关系
        property_value = [0]*len(property_name)
        calculate_formular = match.group(2) + match.group(3)
        query_id = query_info(access_token, file_id, query_family)
        total = len(formular_list) * len(query_id)
        # 对应每条规则公式逐个构件进行属性审查, 不含网页截图约0.24s/条, 含构件快照约1.4s/条
        for element_id in query_id:
            check_property = ''
            size_data = []
            socket_io.emit('select_id', str(element_id))
            try:
                element_info = get_info(access_token, file_id, element_id)
            except:
                print(f"{datetime.datetime.now()}: 构件{element_id}云端数据连接失败！")
                continue
            for i in element_info['data']['properties']:
                if i['group'] == '尺寸标注':
                   size_data = i['items']
            for size_property in size_data:
                if size_property['key'] in property_name:
                    property_value[property_name.index(size_property['key'])] = size_property['value']
            for j in range(0, len(property_name), 1):
                calculate_formular = calculate_formular.replace(property_name[j], str(property_value[j]))
                check_property += property_name[j] + ": " + str(property_value[j]) + "  "
            calculate_result = eval(calculate_formular)
            if calculate_result:
                calculate_result = "符合"
            else:
                calculate_result = "不符合"
            # 将审查结果写入excel审图报告
            element_property = property_filter(element_info, 1)
            check_result = [element_id, element_property['floor'], element_property['family'], element_property['element_name'], check_property, str(calculate_result)]
            print(f"{datetime.datetime.now()} {check_result}")
            for i in range(0, len(check_result), 1):
                worksheet.row_dimensions[row_index].height = 126  # 设置行高
                worksheet.cell(row=row_index, column=i+2).value = check_result[i]
                worksheet.cell(row=row_index, column=i+2).font = Font(name='微软雅黑', size=12, bold=False)
                worksheet.cell(row=row_index, column=i+2).alignment = Alignment(horizontal='center', vertical='center')
                # worksheet.write(row_index, i, label=check_result[i])
            progress = int(row_index*100/total)
            socket_io.emit('temp_progress', str(progress))
            socket_io.sleep(1)  # 服务端发送数据快，客户端接收数据慢，增加延时避免数据黏包
            row_index += 1
    # workbook.save('审图报告.xls')
    time.sleep(2)
    jpg_num = len(os.listdir("snapshots"))
    for i in range(0, jpg_num, 1):
        image = Image(f"snapshots/{i+1}.jpg")
        worksheet.add_image(image, f"A{i+2}")  # 第一张截图是模型整体快照
    print(f"{datetime.datetime.now()}: 当前审图报告已生成！")
    workbook.save('审图报告.xlsx')

# def create_header(sheet):
#     header_name = ['构件ID', '族名称', '构件名称', '审查属性', '审查结果', '规范条文']
#     col_width = [5000, 6000, 6000, 10000, 4000, 10000]
#     style = xlwt.XFStyle()  # 初始化样式
#     font = xlwt.Font()  # 为样式创建字体
#     font.name = '微软雅黑'
#     font.bold = True  # 黑体
#     style.font = font  # 设定样式
#     alignment = xlwt.Alignment()
#     alignment.horz = xlwt.Alignment.HORZ_CENTER  # 居中对齐
#     style.alignment = alignment
#     for i in range(0, len(header_name), 1):
#         sheet.col(i).width = col_width[i]
#         sheet.write(0, i, label=header_name[i], style=style)
#     return sheet

def create_header(worksheet):
    header_name = ['图像快照', '构件ID', '楼层', '族名称', '构件名称', '审查属性', '审查结果', '规范条文']
    header_word = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    col_width = [23, 25, 15, 25, 25, 35, 15, 45]
    # worksheet.row_dimensions[1].height = 126
    for i in range(0, len(header_name), 1):
        worksheet.column_dimensions[header_word[i]].width = col_width[i]  # 设置列宽
        # 行列最小索引值为1
        worksheet.cell(1, i + 1).value = header_name[i]
        worksheet.cell(1, i + 1).font = Font(name='微软雅黑', size=14, bold=True)
        worksheet.cell(1, i + 1).alignment = Alignment(horizontal='center', vertical='center')
    return worksheet
