from flask import Flask, request, send_file, make_response
from flask_cors import CORS
import requests
import base64
import datetime
import json
import re
import jieba.posseg as pseg
import xlrd
from flask_socketio import SocketIO
# from examine_info import query_info, get_info, check_info
import base64
import numpy as np
import cv2
import threading
import time
import os
import shutil
from multiprocessing import Process
import psutil
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
from property_filter import property_filter
from check_materials import check_files

app = Flask(__name__)
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = datetime.timedelta(seconds=1)  # 设置浏览器文件缓存时间为1秒
CORS(app)
socket_io = SocketIO(app, cors_allowed_origins='*')
start_flag = True
pause_flag = False
num = 1
row_index = 2
query_j = 0
formular_j = 0
workbook, worksheet = None, None
result_flag = False
# 获取Access Token(有效期为7天)
access_url = 'https://api.bimface.com/oauth2/token'
access_app = base64.b64encode('9ajg3YC3kcUd8pyXHBv8YtkpLD04bwjh:9wl595y8LbCh30SQK1XOajRTkvg62aWn'.encode())
access_header = {"Authorization": f"Basic {access_app.decode()}"}
access_token = requests.post(url=access_url, headers=access_header).json()['data']['token']
print(f"{datetime.datetime.now()}: 用户验证码为——{access_token}")
# 雨虹110kV变电站项目文件ID
file_id = '2120168448190912'
# 获取模型的View Token(有效期为12h)
@app.route('/view/token')
def get_view_token():
    global result_flag
    if request.method == 'GET':
        result_flag = True
        view_url = 'https://api.bimface.com/view/token'
        view_header = {"Authorization": f"Bearer {access_token}"}
        view_param = {"fileId": file_id}
        view_token = requests.get(url=view_url, headers=view_header, params=view_param).json()['data']
        print(f"{datetime.datetime.now()}: 临时访问凭证推送成功! {view_token}")
        return view_token
# 获取源文件信息
@app.route('/file/info')
def get_file_info():
    if request.method == 'GET':
        file_url = f"https://file.bimface.com/files/{file_id}"
        file_header = {"Authorization": f"Bearer {access_token}"}
        file_info = requests.get(url=file_url, headers=file_header).json()['data']
        print(f"{datetime.datetime.now()}: 项目文件信息推送成功! {file_info}")
        return file_info
# 可研资料完备性审查
@app.route('/check/materials', methods=['GET', 'POST'])
def get_check_materials():
    if request.method == 'POST':
        materials_input = json.loads(request.data)['data']  # {'0':filename0,'1':filename1,...,'N':filenameN}
        print("可研送审资料内容：", materials_input)
        result_files = check_files(materials_input)
        print(f"{datetime.datetime.now()}: 审图报告下载完成！")
        return result_files
# 下载资料检测报告
@app.route('/material/report')
def download_check_materials():
    print(f"{datetime.datetime.now()}: 资料检测报告下载完成！")
    # 设置浏览器URL文件的缓存时间为1s, 及时更新本地文件, 支持断点续传
    return send_file(r'资料检测报告.xlsx', conditional=True, cache_timeout=1)
# 根据elementId获取构件属性(构件ID、楼层、类别ID、类别名称、族、族类型、构件名称、结构材质、体积、三维坐标)
@app.route('/element/property', methods=['GET', 'POST'])  # headers中默认access-allow-methods只有GET
def get_element_property():
    if request.method == 'POST':
        query_input = json.loads(request.data)['data']['query_info']  # 记录请求体中的键值对
        # query_input = request.args['query_info']  # 记录请求URL中的查询参数
        # 输入构件ID(正则匹配结果为纯数字)
        if re.match(r'^\d+$', query_input, flags=0):  # r防止字符转义
            element_id = query_input
            element_num = 1
            element_info = get_info(access_token, file_id, element_id)
            element_property = property_filter(element_info, element_num)
        # 输入构件名称(正则匹配结果为None)
        else:
            query_id = query_info(access_token, file_id, query_input)
            # query_url = "https://api.bimface.com/data/v2/query/elementIds"
            # query_header = {"Authorization": f"Bearer {access_token}"}  # 在请求头中添加上Authorization这个参数
            # query_body = {"targetType": "file", "targetIds": [file_id], "query": {"contain": {"family": query_input}}}  # 查询请求体字典要序列化成JSON字符串
            # query_id = requests.post(url=query_url, headers=query_header, data=json.dumps(query_body)).json()['data'][0]['elementIds']
            element_num = len(query_id)
            if element_num == 0:
                print(f"{datetime.datetime.now()}: {query_input}不存在!")
                return ''
            else:
                element_id = query_id[0]
        # element_url = f"https://api.bimface.com/data/v2/files/{file_id}/elements/{element_id}"
        # element_header = {"Authorization": f"Bearer {access_token}"}
        # element_info = requests.get(url=element_url, headers=element_header).json()
            element_info = get_info(access_token, file_id, element_id)
            element_property = property_filter(element_info, element_num)
            element_property["element_id"] = ", ".join(query_id)
        print(f"构件属性信息为: {element_info}")
        print(f"{datetime.datetime.now()}: 构件属性数据推送成功! {element_property}")
        return element_property
# 回调输入条文结巴分词结果
@app.route('/nlp/clause', methods=['GET', 'POST'])
def get_nlp_segpos():
    if request.method == 'POST':
        clause_input = json.loads(request.data)['data']['clause_input']
        clause_seg = pseg.cut(clause_input)
        seg_result = ''
        nlp_result = {'segpos': '', 'formular': ''}
        for word, pos in clause_seg:
            seg_result += f"{word}/{pos}  "
        data = xlrd.open_workbook("clause.xls")
        table = data.sheets()[0]
        clause_list = [c.value for c in table.col(1, start_rowx=1, end_rowx=table.nrows)]
        formular_list = [c.value for c in table.col(3, start_rowx=1, end_rowx=table.nrows)]
        if clause_input in clause_list:
            nlp_result['segpos'] = seg_result
            nlp_result['formular'] = formular_list[clause_list.index(clause_input)]
        return nlp_result
# 回调审图进度(socket实时同步通讯)
@socket_io.on('check_progress')
def get_check_progress(font_flag):
    global start_flag
    # args的参数只要一个值的时候，参数后面需要加一个逗号，表示args是可迭代的，后面可能还有别的参数，不加逗号会出错
    # socket连接对象只能在线程间传递使用
    # 进程之间默认是不能共享全局变量的, 子进程不能改变主进程中全局变量的值
    # process = Process(target=check_info, args=(access_token, file_id,))
    # pid = process.pid
    # pause = psutil.Process(pid)
    # event = threading.Event()
    # thread = threading.Thread(target=check_info, args=(access_token, file_id, socket_io, event))
    if font_flag == "check_start" and start_flag:
        start_flag = False
        print(f"{datetime.datetime.now()}: Socket实时通讯初始化成功, 开始审查！")
        check_info(access_token, file_id, socket_io)
        # print(f"{datetime.datetime.now()}: 当前图纸审查完毕！")
        # start_flag = True
        # process.start()  # 开始子进程
        # event.set()  # 将内部标识设置为true
        # thread.start()
    # elif font_flag == "check_pause":
    #     pause_flag = True
    #     print(f"{datetime.datetime.now()}: 当前审查已暂停！")
        # pause.suspend()  # 暂停子进程
        # event.clear()  # 将内部标识设置为false
    # elif font_flag == "check_resume":
    #     pause_flag = False
    #     print(f"{datetime.datetime.now()}: 当前审查已恢复！")
        # pause.resume()  # 恢复子进程
        # event.set()
@app.route('/check/status', methods=['GET', 'POST'])
def get_check_status():
    global pause_flag
    if request.method == "POST":
        check_status = json.loads(request.data)['data']['check_status']
        if check_status == "check_pause":
            pause_flag = True
            # print(f"{datetime.datetime.now()}: 当前审查已暂停！")
        elif check_status == "check_resume":
            pause_flag = False
            print(f"{datetime.datetime.now()}: 当前审查已恢复！")
            check_info(access_token, file_id, socket_io, "no_sheet")
    return {"message": "OK"}
# 回调审查结果
@app.route('/check/result')
def get_check_result():
    global result_flag
    data = xlrd.open_workbook("result.xls")
    table = data.sheets()[0]
    result = {"result": []}
    for i in range(1, table.nrows, 1):
        row_value = table.row_values(i, start_colx=0, end_colx=None)
        result['result'].append({'clause':row_value[0], 'content':row_value[1], 'state':row_value[2], 'description':row_value[3], 'id':int(row_value[4])})
    print(f"{datetime.datetime.now()}: 审查结果推送成功！{result}")
    if result_flag:
        return result
    else:
        return {"result": []}
# 下载审查报告
@app.route('/check/report')
def get_check_file():
    if request.method == 'GET':
        # name = f"审图报告-{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S').replace(':', '-')}.xls"
        # os.rename("审图报告.xls", name)
        # shutil.copy(name, "审图报告.xls")
        print(f"{datetime.datetime.now()}: 审图报告下载完成！")
        # 设置浏览器URL文件的缓存时间为1s, 及时更新本地文件, 支持断点续传
        return send_file(r'审图报告(离线版).xlsx', conditional=True, cache_timeout=1)
# 接收审查构件快照截图(960*742->160*160,data:image/png;base64,...), 保持浏览器始终处于当前活动页面
@app.route('/image/snapshot', methods=['GET', 'POST'])
def get_image_snapshot():
    global num
    if request.method == 'POST':
        base64_img = json.loads(request.data)['data']['base64_image'].split(",")[1]
        # base64解码utf-8编码的ASCII文本字节流
        img_data = base64.b64decode(base64_img)
        # 转换为np数组
        img_array = np.frombuffer(img_data, np.uint8)
        # 转换成opencv可用格式
        img_cv2 = cv2.resize(cv2.imdecode(img_array, cv2.COLOR_RGB2BGR), (160, 160))
        cv2.imwrite(f"snapshots/{num}.jpg", img_cv2)
        print(f"{datetime.datetime.now()}: 构件快照{num}接收成功！")
        num += 1
        return {"image_flag": "success"}
# 查询符合条件的构件ID列表
def query_info(access_token, file_id, query_input):
    query_url = "https://api.bimface.com/data/v2/query/elementIds"
    query_header = {"Authorization": f"Bearer {access_token}"}  # 在请求头中添加上Authorization这个参数
    query_body = {"targetType": "file", "targetIds": [file_id],
                  "query": {"contain": {"family": query_input}}}  # 查询请求体字典要序列化成JSON字符串
    query_id = requests.post(url=query_url, headers=query_header, data=json.dumps(query_body)).json()['data'][0]['elementIds']
    return query_id
# 获取指定ID构件的属性
def get_info(access_token, file_id, element_id):
    element_url = f"https://api.bimface.com/data/v2/files/{file_id}/elements/{element_id}"
    element_header = {"Authorization": f"Bearer {access_token}"}
    element_info = requests.get(url=element_url, headers=element_header).json()
    return element_info
# 属性审查模块
def check_info(access_token, file_id, socket_io, sheet_flag=None):
    # event.wait()  # 阻塞线程直到内部变量为true
    global row_index, query_j, formular_j, start_flag, pause_flag, worksheet, workbook
    with open("formular.txt", 'r', encoding='utf-8-sig') as f:  # 非法字符\ufeff
        formular_list = f.read().split("\n")[0:6]
    if sheet_flag != "no_sheet":
        workbook = openpyxl.Workbook()
        worksheet = workbook.create_sheet(title="审查结果", index=0)
        worksheet = create_header(worksheet)
    total = 0
    for k in range(formular_j, len(formular_list), 1):
        formular = formular_list[k]
        match = re.match(r'^\((.*?),(.*?),(.*?)\)(.*?)$', formular)
        query_family = match.group(1)
        query_id = query_info(access_token, file_id, query_family)
        total += len(query_id)
    for k in range(formular_j, len(formular_list), 1):
        query_j = 0
        formular = formular_list[k]
        match = re.match(r'^\((.*?),(.*?),(.*?)\)(.*?)$', formular)
        query_family = match.group(1)
        property_name = match.group(2).split("/")  # 比值关系
        group_name = match.group(3)  # 属性分组数据类型(见预览页属性白色字段)
        property_value = [0]*len(property_name)
        calculate_formular = match.group(2) + match.group(4)
        query_id = query_info(access_token, file_id, query_family)
        # 对应每条规则公式逐个构件进行属性审查, 不含网页截图约0.24s/条, 含构件快照约1.4s/条
        for i in range(query_j, len(query_id), 1):
            if pause_flag is True:
                # 保存上一次暂停时的审查中断状态
                formular_j = k
                query_j = i
                print(f"{datetime.datetime.now()}: 当前审查已暂停！")
                return None
            element_id = query_id[i]
            check_property = ''
            group_data = []
            socket_io.emit('select_id', str(element_id))
            try:
                element_info = get_info(access_token, file_id, element_id)
            except:
                print(f"{datetime.datetime.now()}: 构件{element_id}云端数据连接失败！")
                continue
            for j in element_info['data']['properties']:
                if j['group'] == group_name:
                   group_data = j['items']
            for size_property in group_data:
                if size_property['key'] in property_name:
                    property_value[property_name.index(size_property['key'])] = size_property['value']
            for m in range(0, len(property_name), 1):
                calculate_formular = calculate_formular.replace(property_name[m], str(property_value[m]))
                check_property += property_name[m] + ": " + str(property_value[m]) + "  "
            try:
                calculate_result = eval(calculate_formular)
            except:
                if '乙级' in calculate_formular:
                    calculate_result = eval(calculate_formular.replace('乙级', '2'))
                else:
                    calculate_result = False
            if calculate_result:
                calculate_result = "符合"
            else:
                calculate_result = "不符合"
            # 将审查结果写入excel审图报告
            element_property = property_filter(element_info, 1)
            check_result = [element_id, element_property['floor'], element_property['family'], element_property['element_name'], check_property, str(calculate_result)]
            print(f"{datetime.datetime.now()} {check_result}")
            for n in range(0, len(check_result), 1):
                worksheet.row_dimensions[row_index].height = 126  # 设置行高
                worksheet.cell(row=row_index, column=n+2).value = check_result[n]
                worksheet.cell(row=row_index, column=n+2).font = Font(name='微软雅黑', size=12, bold=False)
                worksheet.cell(row=row_index, column=n+2).alignment = Alignment(horizontal='center', vertical='center')
                # worksheet.write(row_index, i, label=check_result[i])
            progress = int(row_index*100/total)
            socket_io.emit('temp_progress', str(progress))
            # 后端数据审查速度和前端构件加载速度要适配
            socket_io.sleep(2)  # 服务端发送数据快，客户端接收数据慢，增加延时避免数据黏包
            row_index += 1
    # time.sleep(2)
    # jpg_num = len(os.listdir("snapshots"))
    # for i in range(0, jpg_num, 1):
    #     image = Image(f"snapshots/{i+1}.jpg")
    #     worksheet.add_image(image, f"A{i+2}")  # 第一张截图是模型整体快照
    print(f"{datetime.datetime.now()}: 当前审图报告已生成！")
    workbook.save('审图报告.xlsx')
    print(f"{datetime.datetime.now()}: 当前图纸审查完毕！")
    start_flag = True
# 创建审查报告表头
def create_header(worksheet):
    header_name = ['图像快照', '构件ID', '楼层', '族名称', '构件名称', '审查属性', '审查结果', '规范条文']
    header_word = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    col_width = [23, 25, 15, 25, 25, 35, 15, 45]
    # worksheet.row_dimensions[1].height = 126
    for i in range(0, len(header_name), 1):
        worksheet.column_dimensions[header_word[i]].width = col_width[i]  # 设置列宽
        # 行列最小索引值为1
        worksheet.cell(1, i + 1).value = header_name[i]
        worksheet.cell(1, i + 1).font = Font(name='微软雅黑', size=14, bold=True)
        worksheet.cell(1, i + 1).alignment = Alignment(horizontal='center', vertical='center')
    return worksheet

if __name__ == '__main__':
    # app.run(debug=True, port=5000)
    socket_io.run(app, port='5000', debug=True)
