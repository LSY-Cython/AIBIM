import re
import openpyxl
from openpyxl.styles import Font, Alignment
# 审查可研送审资料完备性(文件名模糊匹配)
def check_files(file_dict):
    necessary_files = {"可行性研究报告":[],"本期工程接入系统接线图":[],"站区总体规划图":[],
                       "电气主接线图":[],"电气总平面布置图":[],"土建总平面布置图":[],
                       "配电装置电气平面布置图":[],"线路路径方案图":[],"其他图纸":[]}
    else_files = {"省电力调度中心批复确定的变电站命名文件": [], "电力系统现状地理接线图": [], "电力系统远景规划图": [],
                  "综合楼各层平面布置图": [], "变电站地理位置图": [], "配电装置电气接线图": [],
                  "进出线布置图:": [], "杆塔一览图": [], "基础一览图": [],
                  "金属护套接地方式示意图:": [], "电缆敷设方式图": [], "通信通道组织图": [],
                  "其他文件": []}
    for i, filename in file_dict.items():
        if re.match(r"^(.*?)可(.*?)研(.*?)报告(.*?)$", filename):
            necessary_files["可行性研究报告"].append(filename)
        elif re.match(r"^(.*?)接入(.*?)系统(.*?)图(.*?)$", filename):
            necessary_files["本期工程接入系统接线图"].append(filename)
        elif re.match(r"^(.*?)站区(.*?)总体(.*?)规划(.*?)图(.*?)$", filename):
            necessary_files["站区总体规划图"].append(filename)
        elif re.match(r"^(.*?)电气(.*?)主接线(.*?)图(.*?)$", filename):
            necessary_files["电气主接线图"].append(filename)
        elif re.match(r"^(.*?)电气(.*?)总平面(.*?)布置(.*?)图(.*?)$", filename):
            necessary_files["电气总平面布置图"].append(filename)
        elif re.match(r"^(.*?)土建(.*?)总平面(.*?)布置(.*?)图(.*?)$", filename):
            necessary_files["土建总平面布置图"].append(filename)
        elif re.match(r"^(.*?)电气(.*?)平面(.*?)布置(.*?)图(.*?)$", filename):
            necessary_files["配电装置电气平面布置图"].append(filename)
        elif re.match(r"^(.*?)路径(.*?)方案(.*?)图(.*?)$", filename):
            necessary_files["线路路径方案图"].append(filename)
        elif re.match(r"^(.*?)地理(.*?)接线(.*?)图(.*?)$", filename):
            else_files["电力系统现状地理接线图"].append(filename)
        elif re.match(r"^(.*?)远景(.*?)规划(.*?)图(.*?)$", filename):
            else_files["电力系统远景规划图"].append(filename)
        elif re.match(r"^(.*?)层(.*?)平面(.*?)布置(.*?)图(.*?)$", filename):
            else_files["综合楼各层平面布置图"].append(filename)
        elif re.match(r"^(.*?)地理(.*?)位置(.*?)图(.*?)$", filename):
            else_files["变电站地理位置图"].append(filename)
        elif re.match(r"^(.*?)电气(.*?)接线(.*?)图(.*?)$", filename):
            else_files["配电装置电气接线图"].append(filename)
        elif re.match(r"^(.*?)进出(.*?)线(.*?)布置(.*?)图(.*?)$", filename):
            else_files["进出线布置图"].append(filename)
        elif re.match(r"^(.*?)杆塔(.*?)图(.*?)$", filename):
            else_files["杆塔一览图"].append(filename)
        elif re.match(r"^(.*?)基础(.*?)图(.*?)$", filename):
            else_files["基础一览图"].append(filename)
        elif re.match(r"^(.*?)套(.*?)接地(.*?)图(.*?)$", filename):
            else_files["金属护套接地方式示意图"].append(filename)
        elif re.match(r"^(.*?)电缆(.*?)敷设(.*?)图(.*?)$", filename):
            else_files["电缆敷设方式图"].append(filename)
        elif re.match(r"^(.*?)通信(.*?)通道(.*?)组织(.*?)图(.*?)$", filename):
            else_files["通信通道组织图"].append(filename)
        elif re.match(r"^(.*?)图(.*?)$", filename):
            necessary_files["其他图纸"].append(filename)
        else:
            else_files["其他文件"].append(filename)
    necessary_materials = ''
    option_materials = ''
    for key, file_list in necessary_files.items():
        if file_list == []:
            necessary_files[key] = '未检测到此类文件'
            necessary_materials += f'{key}: 未检测到此类文件\n'
        else:
            necessary_materials += f"{key}: {' '.join(file_list)}\n"
    option_files = []
    for c in list(else_files.values()):
        option_files.extend(c)
    if len(option_files) != 0:
        option_materials += '\n'.join(option_files)
    result_files = {"necessary_materials":necessary_files, "option_materials":option_materials}
    print("可研送审资料解析结果：", result_files)
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(title="完备性验证结果", index=0)
    header_name = ['序号', '资料名称', '检测结果']
    header_word = ['A', 'B', 'C']
    col_width = [25, 45, 60]
    for i in range(0, len(header_name), 1):
        worksheet.column_dimensions[header_word[i]].width = col_width[i]  # 设置列宽
        # 行列最小索引值为1
        worksheet.cell(1, i + 1).value = header_name[i]
        worksheet.cell(1, i + 1).font = Font(name='微软雅黑', size=14, bold=True)
        worksheet.cell(1, i + 1).alignment = Alignment(horizontal='center', vertical='center')
    id = 1
    for key, file in necessary_files.items():
        row_index = id+1
        worksheet.row_dimensions[row_index].height = 50  # 设置行高
        worksheet.cell(row=row_index, column=1).value = str(id)
        worksheet.cell(row=row_index, column=1).font = Font(name='微软雅黑', size=12, bold=False)
        worksheet.cell(row=row_index, column=1).alignment = Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=row_index, column=2).value = key
        worksheet.cell(row=row_index, column=2).font = Font(name='微软雅黑', size=12, bold=False)
        worksheet.cell(row=row_index, column=2).alignment = Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=row_index, column=3).value = "  ".join(file)
        worksheet.cell(row=row_index, column=3).font = Font(name='微软雅黑', size=12, bold=False)
        worksheet.cell(row=row_index, column=3).alignment = Alignment(horizontal='center', vertical='center')
        id += 1
    worksheet.row_dimensions[id+1].height = 150  # 设置行高
    worksheet.cell(row=id + 1, column=1).value = str(id)
    worksheet.cell(row=id + 1, column=1).font = Font(name='微软雅黑', size=12, bold=False)
    worksheet.cell(row=id + 1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=id + 1, column=2).value = "其他文件"
    worksheet.cell(row=id + 1, column=2).font = Font(name='微软雅黑', size=12, bold=False)
    worksheet.cell(row=id + 1, column=2).alignment = Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=id + 1, column=3).value = option_materials
    worksheet.cell(row=id + 1, column=3).font = Font(name='微软雅黑', size=10, bold=False)
    worksheet.cell(row=id + 1, column=3).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    workbook.save('资料检测报告.xlsx')
    print("资料检测报告已生成！")
    return result_files
